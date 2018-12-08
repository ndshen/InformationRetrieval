import os
import numpy as np
import pandas as pd
import math
import class_feature
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice


base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
document_dir = os.path.join(base_dir, "IRTM")
dictionary_file = os.path.join(os.path.dirname(__file__),'dictionary_x2.xlsx')
result_file = os.path.join(os.path.dirname(__file__), 'result_x2.csv')

def load_dictionary(inputFile = dictionary_file) -> tuple:
    """load the dictionary.xlsx into the input format of NB_classifier"""
    wb = load_workbook(filename=inputFile)
    term_df_for_classes = dict()
    dictionary = set()
    sheets = wb.sheetnames
    for sheet in sheets:
        data = wb[sheet].values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        term_df_for_classes[sheet] = df

    dictionary = set(df.index)
    return (dictionary, term_df_for_classes)   

def NB_classifer(doc:str, dictionary:set, classes_df:dict, classes_docs:dict) -> str:
    """A classifier that would determine which class it belongs to"""
    token_list = class_feature.extract_terms_lemma(doc)
    total_d = sum(len(docs) for c, docs in classes_docs.items())
    c_value = dict()
    for c, docs in classes_docs.items():
        c_value[c] = math.log(len(docs)/total_d)

    for token in token_list:
        if token in dictionary:
            for c, df in classes_df.items():
                if df.loc[token, "df_chisq_x"] != 0:
                    c_value[c] += math.log(df.loc[token, "df_chisq_x"])

    return(max(c_value, key= lambda k: c_value[k]))

def test_result():
    result_df = pd.read_csv(result_file)
    classes_doc = class_feature.get_training_classes()
    err = 0

    for c, docs in classes_doc.items():
        for doc in docs:
            if int(result_df.loc[int(doc)-1, "value"]) != int(c):
                print(doc)
                err += 1
            break
    
    print("error num: {}".format(err))
    print("error rate: {}".format(err/sum(len(docs) for c, docs in classes_doc.items())))

if __name__ == "__main__":
    dictionary, classes_df = load_dictionary()
    testing_list = os.listdir(document_dir)
    classes_doc = class_feature.get_training_classes()
    training_data = set()
    for c, docs in classes_doc.items():
        training_data = training_data.union(set(docs))

    result_dict = dict()
    for i, doc in enumerate(testing_list):
        if doc[:-4] in training_data:
            continue
        text = class_feature.fetch_document(doc[:-4])
        c = NB_classifer(text, dictionary, classes_df, classes_doc)
        result_dict[int(doc[:-4])] = c
        print(i, end='\r')
    
    result_pd = pd.DataFrame(list(result_dict.items()), columns=["Id", "Value"])
    result_pd = result_pd.sort_values(by=["Id"], ascending=True)
    result_pd.to_csv(result_file, index=0)
    
    # test_result()