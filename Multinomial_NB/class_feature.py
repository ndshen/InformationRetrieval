import os
import re
from collections import defaultdict
import numpy as np
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice

import nltk
from nltk import sent_tokenize, pos_tag
from nltk.corpus import stopwords, wordnet
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
document_dir = os.path.join(base_dir, "IRTM")
training_input = os.path.join(os.path.dirname(__file__), "training.txt")
tf_df_file = os.path.join(os.path.dirname(__file__),'tf_df.xlsx')
feature_file =  os.path.join(os.path.dirname(__file__),'class_feature_x2.xlsx')
dictionary_file = os.path.join(os.path.dirname(__file__),'dictionary_x2.xlsx')

def fetch_document(id:int) -> str:
    with open(os.path.join(document_dir, "{}.txt".format(id)), "r") as f:
        text = f.read()
    return(text)

def extract_terms_lemma(text_file):
    """Uses NLTK Lemmatizer module to extract token from a text file"""

    def get_wordnet_pos(treebank_tag):
        """Get the wordnet pos for WordNetLemmatizer"""
        if treebank_tag.startswith('J'):
            return wordnet.ADJ
        elif treebank_tag.startswith('V'):
            return wordnet.VERB
        elif treebank_tag.startswith('N'):
            return wordnet.NOUN
        elif treebank_tag.startswith('R'):
            return wordnet.ADV
        else:
            return None
    
    sentences = sent_tokenize(text_file) #split text_file to list of sentences

    tokens = []
    lemmatizer = WordNetLemmatizer()
    for sentence in sentences:
        for word, pos in pos_tag(word_tokenize(sentence)):
            wordnet_pos = get_wordnet_pos(pos) or wordnet.NOUN
            tokens.append(lemmatizer.lemmatize(word, pos=wordnet_pos))

    # As long as there is a non-punctuation exist in the word, keep it.
    # It allow words like 'Mr.' or 'run-off' to be in the final token list
    words = [word.lower() for word in tokens if re.search(r'\w+', word)] #filter out punctuation and lower cases

    stop_words = set(stopwords.words('english'))
    custom_stop_words = {'\'s', '\'t', '\'ll', '\'m', 'n\'t'} # ex: cat's, don't
    stop_words = stop_words.union(custom_stop_words) #add custom stop words
    final_token_list = [word for word in words if word not in stop_words and len(word) != 1] #filter out stop words and single letter
    
    return(final_token_list)

def get_training_classes(inputFile=training_input) -> dict:
    """Convert the training data into dictionary type"""
    training_classes = dict()
    with open(inputFile, "r") as f:
        lines = f.read().split('\n')
        for line in lines:
            entry = line.split(' ')
            c = entry[0]
            docs = [doc for doc in entry[1:] if doc != '']
            training_classes[c] = docs
        
    return(training_classes)

def calculate_tf_df(training_classes):
    whole_dictionary = defaultdict(dict)
    term_tables_for_classes = defaultdict(dict)
    for c, docs in training_classes.items():
        print("Start to process class {} ==============================".format(c))
        term_table = defaultdict(dict)
        for doc in docs:
            print("Start to process document {}".format(doc))
            tokens = extract_terms_lemma(fetch_document(int(doc)))
            token_set = set()
            for token in tokens:
                if token in token_set:
                    whole_dictionary[token][0] += 1
                    term_table[token][0] += 1
                else:
                    token_set.add(token)
                    if token in term_table:
                        whole_dictionary[token][0] += 1
                        whole_dictionary[token][1] += 1
                        term_table[token][0] += 1
                        term_table[token][1] += 1
                    else:
                        term_table[token] = [1, 1]
                        if token in whole_dictionary:
                            whole_dictionary[token][0] += 1
                            whole_dictionary[token][1] += 1
                        else:
                            whole_dictionary[token] = [1, 1]
        
        term_tables_for_classes[c] = pd.DataFrame.from_dict(term_table, orient='index',columns=['tf', 'df'])

    output_file = pd.ExcelWriter(tf_df_file)
    main_term_table = pd.DataFrame.from_dict(whole_dictionary, orient='index',columns=['tf', 'df'])
    main_term_table.to_excel(output_file, "all")
    for c, df in term_tables_for_classes.items():
        df.to_excel(output_file, c)
    output_file.save()
    print(main_term_table)

def load_tf_df_file(inputFile = tf_df_file) -> tuple:
    """load the tf_df.xlsx file into a main dataframe(all) and a list of class dataframes"""
    wb = load_workbook(filename=inputFile)
    term_df_for_classes = dict()
    sheets = wb.sheetnames
    for sheet in sheets:
        data = wb[sheet].values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        if sheet == "all":
            main_term_df = df
        else:
            term_df_for_classes[sheet] = df

    return (main_term_df, term_df_for_classes)

def feature_select(all_df, classes_df:dict, classes_docs:dict, size=500, method="df_chisq", outputFile=feature_file):
    
    output_file = pd.ExcelWriter(outputFile)
    total_d = sum((len(docs) for c, docs in classes_docs.items()))
    
    def df_chisq(c:str, df):

        class_d = len(classes_docs[c])
        df["expected_df"] = round(df["df_all"] * (class_d/total_d), 3)
        df["df_chisq"] = round(pow(df["df_class"] - df["expected_df"], 2)/df["expected_df"], 3)
        df['df_chisq_pn'] = df.apply(lambda row: row["df_chisq"] * -1 if row["expected_df"] > row["df_class"] else row["df_chisq"], axis=1)
        df['df_chisq_x'] = df.apply(lambda row:math.log2(row["tf_class"] + 1) * row["df_chisq_pn"], axis=1) #custome adjust on df_chisq value
        df = df.sort_values(by=["df_chisq_x"], ascending=False)
        return(df.head(size))

    for c, df in classes_df.items():
        class_df_merged = pd.merge(df, all_df, left_index=True, right_index=True, suffixes=['_class', '_all'], how='left')
        _locals = locals()
        exec("new_class_df = {}(\"{}\", class_df_merged)".format(method, c), globals(), _locals) #allow different feature selection method to be implemented by argument
        _locals["new_class_df"].to_excel(output_file, c)
    
    output_file.save()

def load_class_feature_file(inputFile = feature_file) -> dict:
    """load the class_feature.xlsx into a the input format of construct_dictionary"""
    wb = load_workbook(filename=inputFile)
    term_df_for_classes = dict()
    sheets = wb.sheetnames
    for sheet in sheets:
        data = wb[sheet].values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        term_df_for_classes[sheet] = df

    return (term_df_for_classes)    

def construct_dictionary(classes_df:dict, classes_docs:dict, size=500, outputFile=dictionary_file):
    """
    Consturct a dictionary from class feature by choosing only 500 terms,
    and calculate the probability for each term in each dictionary.
    """
    output_file = pd.ExcelWriter(outputFile)
    def select_terms_strict():
        """select top 500 df_chisq_x terms, cause the dictionary size to be strict 500."""
        whole_pd = pd.concat([df for c, df in classes_df.items()], axis=0)
        whole_pd = whole_pd.sort_values(by=["df_chisq_x"], ascending=False)
        whole_pd = whole_pd[~whole_pd.index.duplicated(keep='first')] #delete duplicate terms, and keep the highest df_chisq_x value
        dictionary = whole_pd[["tf_all", "df_all"]].head(size)
        return(dictionary)
    
    dictionary = select_terms_strict()
    for c, df in classes_df.items():
        c_dict_merged = pd.merge(dictionary, df[["tf_class", "df_class", "df_chisq_x"]], left_index=True, right_index=True, how="left")
        c_dict_merged = c_dict_merged.fillna(0).sort_values(by=["tf_class"], ascending=False)
        tf_sum = c_dict_merged["tf_class"].sum()
        c_dict_merged["prob"] = c_dict_merged.apply(lambda row:(row["tf_class"]+1)/(tf_sum+size), axis=1) #add-one smoothing
        c_dict_merged.to_excel(output_file, c)
    
    output_file.save()

if __name__ == "__main__":
    # calculate_tf_df(get_training_classes())
    all_df, classes_df = load_tf_df_file()
    feature_select(all_df, classes_df, get_training_classes())
    construct_dictionary(load_class_feature_file(), get_training_classes())
